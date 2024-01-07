import openpyxl as op 
from openpyxl.chart import BarChart, Reference

wb = op.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2,sheet.max_row+1):
    original_value = sheet.cell(row,3).value
    correct_value = original_value * 0.9
    sheet.cell(row,4).value = correct_value

values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')
wb.save('transactions2.xlsx')